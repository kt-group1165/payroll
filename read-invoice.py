from openpyxl import load_workbook
import sys
sys.stdout.reconfigure(encoding="utf-8")

path = r"C:/Users/domen-PC/Desktop/給与計算ソフト/請求管理機能/20250825_介護請求書ひな形_moreの見出しを少し変えた.xlsx"
wb = load_workbook(path, data_only=True)
print("Sheet names:", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n=== Sheet: {name} (rows={ws.max_row}, cols={ws.max_column}) ===")
    # Print first 60 rows × all cols with non-empty values
    for row in range(1, min(ws.max_row + 1, 80)):
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=row, column=col).value
            if v not in (None, ""):
                print(f"  ({row},{col}) = {repr(v)}")
